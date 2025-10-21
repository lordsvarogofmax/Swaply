import os
import logging
import asyncio
import sqlite3
import json
from flask import Flask, request, send_file
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)
import httpx
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from datetime import datetime, timedelta
from selectolax.parser import HTMLParser
from dotenv import load_dotenv
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import tempfile

# === Настройки ===
load_dotenv()
BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")
MODEL = os.environ.get("OPENROUTER_MODEL", "meta-llama/llama-3.1-70b-instruct")
ADMIN_ID = int(os.environ.get("ADMIN_ID", "364191893"))

if not BOT_TOKEN or not OPENROUTER_API_KEY:
    error_msg = "❌ ОШИБКА КОНФИГУРАЦИИ:\n"
    if not BOT_TOKEN:
        error_msg += "• BOT_TOKEN не задан\n"
    if not OPENROUTER_API_KEY:
        error_msg += "• OPENROUTER_API_KEY не задан\n"
    error_msg += "\nСоздайте файл .env с переменными:\n"
    error_msg += "BOT_TOKEN=ваш_токен_от_botfather\n"
    error_msg += "OPENROUTER_API_KEY=ваш_ключ_openrouter\n"
    error_msg += "ADMIN_ID=364191893\n"
    error_msg += "PORT=10000"
    
    print(error_msg)
    logging.error(error_msg)
    exit(1)

# === База данных ===
def init_database():
    conn = sqlite3.connect('bot_feedback.db')
    cursor = conn.cursor()
    
    # Таблица для отслеживания взаимодействий пользователей
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_interactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            username TEXT,
            first_name TEXT,
            last_name TEXT,
            question TEXT NOT NULL,
            answer TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            session_id TEXT,
            feedback_given BOOLEAN DEFAULT FALSE
        )
    ''')
    
    # Таблица для обратной связи
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            interaction_id INTEGER,
            rating INTEGER NOT NULL CHECK (rating >= 1 AND rating <= 5),
            comment TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (interaction_id) REFERENCES user_interactions (id)
        )
    ''')
    
    conn.commit()
    conn.close()

# Инициализируем базу данных
init_database()

# === Вспомогательные функции для работы с БД ===
def save_interaction(user_id, username, first_name, last_name, question, answer, session_id=None):
    conn = sqlite3.connect('bot_feedback.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO user_interactions (user_id, username, first_name, last_name, question, answer, session_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (user_id, username, first_name, last_name, question, answer, session_id))
    interaction_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return interaction_id

def save_feedback(user_id, interaction_id, rating, comment):
    conn = sqlite3.connect('bot_feedback.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO feedback (user_id, interaction_id, rating, comment)
        VALUES (?, ?, ?, ?)
    ''', (user_id, interaction_id, rating, comment))
    
    # Отмечаем, что обратная связь была дана
    cursor.execute('''
        UPDATE user_interactions SET feedback_given = TRUE WHERE id = ?
    ''', (interaction_id,))
    
    conn.commit()
    conn.close()

def get_user_interaction_count(user_id, days=30):
    conn = sqlite3.connect('bot_feedback.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT COUNT(*) FROM user_interactions 
        WHERE user_id = ? AND timestamp >= datetime('now', '-{} days')
    '''.format(days), (user_id,))
    count = cursor.fetchone()[0]
    conn.close()
    return count

def has_given_feedback(user_id, interaction_id):
    conn = sqlite3.connect('bot_feedback.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT COUNT(*) FROM feedback WHERE user_id = ? AND interaction_id = ?
    ''', (user_id, interaction_id))
    count = cursor.fetchone()[0]
    conn.close()
    return count > 0

def get_admin_stats(days=30):
    conn = sqlite3.connect('bot_feedback.db')
    
    # Получаем статистику за последние N дней
    query = '''
        SELECT 
            ui.user_id,
            ui.username,
            ui.first_name,
            ui.last_name,
            ui.question,
            ui.answer,
            ui.timestamp,
            f.rating,
            f.comment
        FROM user_interactions ui
        LEFT JOIN feedback f ON ui.id = f.interaction_id
        WHERE ui.timestamp >= datetime('now', '-{} days')
        ORDER BY ui.timestamp DESC
    '''.format(days)
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df

# === Функции для работы с историей диалога ===
def add_to_conversation_history(user_data, question, answer):
    """Добавляет вопрос и ответ в историю диалога пользователя"""
    if 'conversation_history' not in user_data:
        user_data['conversation_history'] = []
    
    # Добавляем новую пару вопрос-ответ
    user_data['conversation_history'].append({
        'question': question,
        'answer': answer,
        'timestamp': datetime.now().isoformat()
    })
    
    # Ограничиваем историю 10 парами (20 сообщений)
    if len(user_data['conversation_history']) > 10:
        user_data['conversation_history'] = user_data['conversation_history'][-10:]
    
    return user_data['conversation_history']

def get_conversation_context(user_data):
    """Возвращает контекст диалога для передачи в ИИ"""
    if 'conversation_history' not in user_data or not user_data['conversation_history']:
        return ""
    
    context_parts = []
    for i, entry in enumerate(user_data['conversation_history'], 1):
        context_parts.append(f"Вопрос {i}: {entry['question']}")
        context_parts.append(f"Ответ {i}: {entry['answer']}")
    
    return "\n\n".join(context_parts)

def clear_conversation_history(user_data):
    """Очищает историю диалога пользователя"""
    user_data['conversation_history'] = []
    return user_data

# === Логирование ===
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# === Загрузка базы знаний из base_knowledge/*.txt ===
_knowledge_chunks = []
knowledge_dir = "base_knowledge"

if os.path.exists(knowledge_dir):
    txt_files = sorted(glob.glob(os.path.join(knowledge_dir, "*.txt")))
    if not txt_files:
        print("⚠️ В папке base_knowledge нет .txt файлов")
    for path in txt_files:
        try:
            with open(path, "r", encoding="utf-8") as f:
                text = f.read().strip()
                # Разбиваем на чанки по ~500 слов
                sentences = text.split(". ")
                current_chunk = ""
                for sent in sentences:
                    if len(current_chunk.split()) + len(sent.split()) > 500:
                        if len(current_chunk) > 50:
                            _knowledge_chunks.append(current_chunk.strip())
                        current_chunk = sent + ". "
                    else:
                        current_chunk += sent + ". "
                if current_chunk and len(current_chunk) > 50:
                    _knowledge_chunks.append(current_chunk.strip())
        except Exception as e:
            logging.error(f"Ошибка чтения {path}: {e}")

    print(f"✅ Загружено {len(_knowledge_chunks)} фрагментов из base_knowledge/")
else:
    print("❌ Папка base_knowledge не найдена!")

# Подготавливаем TF-IDF векторизатор
if _knowledge_chunks:
    _vectorizer = TfidfVectorizer(stop_words=None)
    _vectorizer.fit(_knowledge_chunks)
    _knowledge_ready = True
else:
    _knowledge_ready = False

# === Приветствие ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    welcome_text = (
        "👷‍♂️ **Я — профессиональный консультант в сфере строительства и ремонта.**\n\n"
        "Специализируюсь на:\n"
        "• Ремонте квартир и домов\n"
        "• Отделке (штукатурка, покраска, плитка)\n"
        "• Электромонтаже и сантехнике\n"
        "• Выборе материалов и инструментов\n"
        "• Расчёте смет и сроков\n\n"
        "• Отвечу на вопросы по СНиПам и ГОСТам\n\n"
        "💬 **Система диалога**: Я помню контекст наших разговоров и могу отвечать на уточняющие вопросы. "
        "Поддерживаю историю диалога для более точных ответов.\n\n"
        "⚠️ **Примечание**: я отвечаю **только по теме строительства и ремонта**.\n\n"
        "Готов помочь вам с реальной задачей — просто опишите её."
    )
    keyboard = [[InlineKeyboardButton("💬 Получить новую консультацию", callback_data="ask")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(welcome_text, parse_mode="Markdown", reply_markup=reply_markup)

# === Обработка нажатия кнопки ===
async def ask_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    # Проверяем, есть ли история диалога
    conversation_history = context.user_data.get('conversation_history', [])
    history_count = len(conversation_history)
    
    if history_count >= 10:
        # Если достигли лимита в 10 пар, очищаем историю
        clear_conversation_history(context.user_data)
        await query.edit_message_text(
            "🔄 **История диалога очищена**\n\n"
            "📝 Напишите ваш новый вопрос по строительству или ремонту. Например:\n\n"
            "• Как выровнять стены гипсокартоном?\n"
            "• Нужна ли гидроизоляция в ванной под плитку?\n"
            "• Какой краской покрасить деревянный пол?\n\n"
            "Я дам развернутый, профессиональный ответ на основе строительных норм и справочников."
        )
    else:
        await query.edit_message_text(
            "📝 Напишите ваш вопрос по строительству или ремонту. Например:\n\n"
            "• Как выровнять стены гипсокартоном?\n"
            "• Нужна ли гидроизоляция в ванной под плитку?\n"
            "• Какой краской покрасить деревянный пол?\n\n"
            "Я дам развернутый, профессиональный ответ на основе строительных норм и справочников."
        )
    
    context.user_data["in_consultation"] = True

# === Поиск релевантных фрагментов ===
def retrieve_relevant_chunks(query, top_k=3):
    if not _knowledge_ready:
        return []
    query_vec = _vectorizer.transform([query])
    knowledge_vecs = _vectorizer.transform(_knowledge_chunks)
    similarities = cosine_similarity(query_vec, knowledge_vecs).flatten()
    top_indices = similarities.argsort()[-top_k:][::-1]
    return [_knowledge_chunks[i] for i in top_indices if similarities[i] > 0.1]

# === Ищет актуальные нормативные документы на docs.cntd.ru ===

async def search_cntd(query: str, max_chars=1500) -> str:
    try:
        async with httpx.AsyncClient(
            timeout=10.0,
            follow_redirects=True,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
                "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Sec-Fetch-User": "?1",
            }
        ) as client:
            search_url = f"https://docs.cntd.ru/search?text={query}"
            response = await client.get(search_url)
            response.raise_for_status()

            # Парсинг без изменений
            tree = HTMLParser(response.text)
            results = []
            for item in tree.css("div.search-results__item"):
                title_node = item.css_first("a")
                if not title_node:
                    continue
                title = title_node.text().strip()
                href = title_node.attributes.get("href")
                if not href or not href.startswith("/document/"):
                    continue
                status_node = item.css_first("span.document-info__status")
                status = status_node.text().strip().lower() if status_node else ""
                if "отмен" in status or "не действует" in status:
                    continue
                results.append({"title": title, "url": "https://docs.cntd.ru" + href})

            if not results:
                return ""

            doc_url = results[0]["url"]
            doc_resp = await client.get(doc_url)
            doc_tree = HTMLParser(doc_resp.text)
            content = ""
            for p in doc_tree.css("div.document-content p"):
                text = p.text().strip()
                if text and len(text) > 20:
                    content += text + "\n"
                    if len(content) > max_chars:
                        break

            return f"[Источник: {results[0]['title']}]\n{content[:max_chars]}..." if content else ""

    except Exception as e:
        logging.error(f"Ошибка поиска на cntd.ru: {e}")
        return ""
        
# === Обработка обратной связи ===
async def handle_feedback_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not query.data.startswith("rating_"):
        return
    
    rating = int(query.data.split("_")[1])
    interaction_id = context.user_data.get("current_interaction_id")
    
    if not interaction_id:
        await query.edit_message_text("❌ Ошибка: не найден ID взаимодействия")
        return
    
    # Сохраняем оценку
    save_feedback(update.effective_user.id, interaction_id, rating, None)
    
    await query.edit_message_text(
        f"✅ Спасибо за оценку {rating} звезд! "
        "Теперь вы можете оставить комментарий или задать новый вопрос.",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("💬 Задать новый вопрос", callback_data="ask")],
            [InlineKeyboardButton("📝 Оставить комментарий", callback_data="comment")]
        ])
    )

async def handle_comment_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "📝 Пожалуйста, напишите ваш комментарий к оценке. "
        "Ваше мнение поможет улучшить качество консультаций."
    )
    
    # Устанавливаем состояние ожидания комментария
    context.user_data["waiting_for_comment"] = True

async def handle_feedback_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("waiting_for_comment"):
        return
    
    comment = update.message.text.strip()
    interaction_id = context.user_data.get("current_interaction_id")
    
    if not interaction_id:
        return
    
    # Обновляем комментарий в базе
    conn = sqlite3.connect('bot_feedback.db')
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE feedback SET comment = ? WHERE interaction_id = ? AND user_id = ?
    ''', (comment, interaction_id, update.effective_user.id))
    conn.commit()
    conn.close()
    
    await update.message.reply_text(
        "✅ Спасибо за комментарий! Ваше мнение поможет улучшить качество консультаций.",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("💬 Задать новый вопрос", callback_data="ask")
        ]])
    )
    
    # Очищаем состояние
    context.user_data.pop("current_interaction_id", None)
    context.user_data.pop("waiting_for_comment", None)

# === Админские команды ===
async def handle_admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ У вас нет прав для выполнения этой команды.")
        return
    
    try:
        # Получаем статистику за последние 30 дней
        df = get_admin_stats(30)
        
        if df.empty:
            await update.message.reply_text("📊 За последние 30 дней нет данных для анализа.")
            return
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика бота"
        
        # Заголовки
        headers = [
            "ID пользователя", "Имя пользователя", "Имя", "Фамилия", 
            "Вопрос", "Ответ", "Дата/время", "Оценка", "Комментарий"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Данные
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row['user_id'])
            ws.cell(row=row_idx, column=2, value=row['username'] or '')
            ws.cell(row=row_idx, column=3, value=row['first_name'] or '')
            ws.cell(row=row_idx, column=4, value=row['last_name'] or '')
            ws.cell(row=row_idx, column=5, value=row['question'][:100] + '...' if len(str(row['question'])) > 100 else row['question'])
            ws.cell(row=row_idx, column=6, value=row['answer'][:100] + '...' if len(str(row['answer'])) > 100 else row['answer'])
            ws.cell(row=row_idx, column=7, value=row['timestamp'])
            ws.cell(row=row_idx, column=8, value=row['rating'] if pd.notna(row['rating']) else 'Нет оценки')
            ws.cell(row=row_idx, column=9, value=row['comment'] or '')
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Отправляем файл
        with open(tmp_file_path, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=f"bot_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                caption=f"📊 Статистика использования бота за последние 30 дней\n"
                       f"Всего взаимодействий: {len(df)}\n"
                       f"Уникальных пользователей: {df['user_id'].nunique()}\n"
                       f"Средняя оценка: {df['rating'].mean():.2f}" if not df['rating'].isna().all() else "Оценок пока нет"
            )
        
        # Удаляем временный файл
        os.unlink(tmp_file_path)
        
    except Exception as e:
        logging.error(f"Ошибка при создании статистики: {e}")
        await update.message.reply_text("❌ Ошибка при создании статистики. Проверьте логи.")

# === Обработка текстовых сообщений ===
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Проверяем, ждем ли мы комментарий
    if context.user_data.get("waiting_for_comment"):
        await handle_feedback_comment(update, context)
        return
    
    if not context.user_data.get("in_consultation", False):
        await start(update, context)
        return

    user_text = update.message.text.strip()

    # Проверка: запрос про нормативы?
    is_normative = any(word in user_text.lower() for word in [
        "снип", "гост", "сп ", "свод правил", "актуальн", "действует",
        "новый", "нормы", "требован", "стандарт", "2025", "2024", "обновл"
    ])

    relevant_chunks = []
    online_context = ""

    if is_normative:
        online_context = await search_cntd(user_text)
        if not online_context:
            # Fallback на локальную базу
            relevant_chunks = retrieve_relevant_chunks(user_text)
    else:
        relevant_chunks = retrieve_relevant_chunks(user_text)

    # Формируем контекст
    if online_context:
        knowledge_context = online_context
    elif relevant_chunks:
        knowledge_context = "\n\n".join(relevant_chunks)
    else:
        knowledge_context = "Нет релевантной информации в базе знаний."

    # Получаем контекст диалога
    conversation_context = get_conversation_context(context.user_data)
    
    # Системный промпт с указанием года
    current_year = datetime.now().year
    system_prompt = (
        f"Сегодня {current_year} год. Ты — профессиональный строитель с 10-летним опытом. "
        "Ты отвечаешь ТОЛЬКО на вопросы по строительству и ремонту. "
        "Если вопрос не по теме — вежливо откажись. "
        "ОТВЕЧАЙ ТОЛЬКО НА РУССКОМ ЯЗЫКЕ. "
        "НЕ ПИШИ РАССУЖДЕНИЯ. НЕ ИСПОЛЬЗУЙ ФРАЗЫ ВРОДЕ «Я думаю». "
        "ПИШИ КАК ЭКСПЕРТ: КРАТКО, ТОЧНО, ПО ДЕЛУ. "
        "МАКСИМУМ — 400 слов."
    )

    # Формируем промпт с учетом истории диалога
    if conversation_context:
        conversation_part = f"Предыдущий диалог с клиентом:\n{conversation_context}\n\n"
    else:
        conversation_part = ""

    if online_context or relevant_chunks:
        user_prompt = (
            f"{conversation_part}"
            f"Информация из строительных нормативов:\n{knowledge_context}\n\n"
            f"Текущий вопрос клиента: {user_text}\n\n"
            f"Ответь на русском языке, без лишних слов."
        )
    else:
        user_prompt = f"{conversation_part}Текущий вопрос клиента: {user_text}\n\nОтветь на русском языке, без лишних слов."

    await update.message.reply_text("⏳ Минутку, мне нужно подумать...")
    logging.info(f"Отправляю запрос к OpenRouter: {user_prompt[:200]}...")

    try:
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]
        last_error = None
        for attempt in range(3):
            try:
                async with httpx.AsyncClient(timeout=30.0) as client:
                    response = await client.post(
                        "https://openrouter.ai/api/v1/chat/completions",
                        headers={
                            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                            "Content-Type": "application/json"
                        },
                        json={
                            "model": MODEL,
                            "messages": messages,
                            "max_tokens": 1000,
                            "temperature": 0.3
                        }
                    )
                if response.status_code == 200:
                    data = response.json()
                    answer = data["choices"][0]["message"]["content"].strip()
                    logging.info(f"Получен ответ от OpenRouter: {answer[:200]}")

                    # Сохраняем взаимодействие в БД
                    user = update.effective_user
                    interaction_id = save_interaction(
                        user.id, 
                        user.username, 
                        user.first_name, 
                        user.last_name, 
                        user_text, 
                        answer
                    )
                    
                    # Добавляем в историю диалога
                    add_to_conversation_history(context.user_data, user_text, answer)
                    
                    # Проверяем, нужно ли запросить обратную связь (после 3-го взаимодействия)
                    interaction_count = get_user_interaction_count(user.id)
                    
                    if interaction_count == 3 and not has_given_feedback(user.id, interaction_id):
                        # Показываем запрос на оценку
                        keyboard = [
                            [InlineKeyboardButton("⭐", callback_data="rating_1")]
                            [InlineKeyboardButton("⭐⭐", callback_data="rating_2")]
                            [InlineKeyboardButton("⭐⭐⭐", callback_data="rating_3")],
                            [InlineKeyboardButton("⭐⭐⭐⭐", callback_data="rating_4")]
                            [InlineKeyboardButton("⭐⭐⭐⭐⭐", callback_data="rating_5")]
                        ]
                        reply_markup = InlineKeyboardMarkup(keyboard)
                        
                        await update.message.reply_text(
                            f"{answer}\n\n"
                            "📊 **Пожалуйста, оцените качество ответа:**\n"
                            "Насколько полезной была информация? (1-5 звезд)",
                            reply_markup=reply_markup,
                            disable_web_page_preview=True,
                            parse_mode="Markdown"
                        )
                        
                        # Сохраняем ID взаимодействия для обратной связи
                        context.user_data["current_interaction_id"] = interaction_id
                    else:
                        # Обычный ответ с кнопкой нового вопроса
                        conversation_history = context.user_data.get('conversation_history', [])
                        history_count = len(conversation_history)
                        remaining = 10 - history_count
                        
                        keyboard = [[InlineKeyboardButton("💬 Задать новый вопрос", callback_data="ask")]]
                        reply_markup = InlineKeyboardMarkup(keyboard)
                        await update.message.reply_text(
                            answer,
                            reply_markup=reply_markup,
                            disable_web_page_preview=True
                        )
                    break
                else:
                    last_error = f"HTTP {response.status_code}: {response.text}"
                    logging.warning(f"Попытка {attempt+1}/3 не удалась: {last_error}")
            except Exception as inner_e:
                last_error = str(inner_e)
                logging.warning(f"Попытка {attempt+1}/3 завершилась ошибкой: {last_error}")
            await asyncio.sleep(1.5 * (attempt + 1))

        else:
            raise Exception(last_error or "Неизвестная ошибка при обращении к OpenRouter")

    except Exception as e:
        logging.error(f"Ошибка ИИ: {e}")
        await update.message.reply_text(
            "⚠️ Не удалось получить ответ. Возможно, временные проблемы с сервисом. "
            "Попробуйте через минуту или переформулируйте вопрос."
        )

# === Flask + Webhook ===
app = Flask(__name__)

application = Application.builder().token(BOT_TOKEN).build()

application.add_handler(CommandHandler("start", start))
application.add_handler(CommandHandler("stats", handle_admin_stats))
application.add_handler(CallbackQueryHandler(ask_callback, pattern="^ask$"))
application.add_handler(CallbackQueryHandler(handle_feedback_rating, pattern="^rating_"))
application.add_handler(CallbackQueryHandler(handle_comment_callback, pattern="^comment$"))
application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

_loop = asyncio.new_event_loop()
asyncio.set_event_loop(_loop)
_loop.run_until_complete(application.initialize())

@app.route("/<path:token>", methods=["POST"])
def telegram_webhook(token):
    if token != BOT_TOKEN:
        return "Forbidden", 403
    data = request.get_json(force=True)
    update = Update.de_json(data, application.bot)
    _loop.run_until_complete(application.process_update(update))
    return "OK"

@app.route("/health")
def health():
    return "OK"

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
